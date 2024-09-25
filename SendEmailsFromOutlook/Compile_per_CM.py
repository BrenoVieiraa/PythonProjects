import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle, PatternFill
import win32com.client as win32
import datetime

list_materials = []
list_per_supplier = []
list_emails=[]
workbook = Workbook()
outlook = win32.Dispatch('outlook.application')



class list:

    def list_of_pending(self):
        
        book = openpyxl.load_workbook(r'C:\Users\lucas.sampaio\PACCAR Inc\Purchasing Brasil-DAF-Teams - Documentos\NPG - Brasil\2 - Macro\1 - Email\Macro - Email - Envio de cotação.xlsx',data_only=True)
        Dados_page = book['Envio']

        count_of_rows = 0

        for rows in Dados_page.iter_rows(min_row=3):
            material = rows[1].value
            descricao = rows[2].value
            quantidade = rows[3].value
            purchase_request = rows[4].value
            fornecedor_1 = rows[5].value
            email_1 = rows[6].value
            fornecedor_2 = rows[7].value
            email_2 = rows[8].value
            fornecedor_3 = rows[9].value
            email_3 = rows[10].value

            info_line = {
                "Material":material,
                "Descrição":descricao,
                "Quantidade":quantidade,
                "Purchase Request": purchase_request,
                "Fornecedor 1": fornecedor_1,
                "Email 1": email_1,
                "Fornecedor 2": fornecedor_2,
                "Email 2": email_2,
                "Fornecedor 3": fornecedor_3,
                "Email 3": email_3
            }

            list_materials.append(info_line)

            count_of_rows += 1

        print("Total: {count_of_rows}")
        print(list_materials)

    def list_per_cm(self):

        for line in list_materials:
            if line["Fornecedor 1"] not in list_per_supplier:
                list_per_supplier.append(line["Fornecedor 1"])

        for line in list_materials:
            if line["Fornecedor 2"] not in list_per_supplier:
                list_per_supplier.append(line["Fornecedor 2"])

        for line in list_materials:
            if line["Fornecedor 3"] not in list_per_supplier:
                list_per_supplier.append(line["Fornecedor 3"])

        for supplier in list_per_supplier:
            workbook = openpyxl.Workbook()
            information = workbook.active

            # Insira os dados da tabela
            information.append(["Material",
                                "Descrição",
                                "Quantidade",
                                "Purchase Request"])

            for line in list_materials:
                if line["Fornecedor 1"] == supplier or line["Fornecedor 2"] == supplier or line["Fornecedor 3"] == supplier:
                    information.append([line["Material"],
                                        line["Descrição"],
                                        line["Quantidade"],
                                        line["Purchase Request"],
                                        ])

            # Ajuste o tamanho das colunas com base no conteúdo
            for column_cells in information.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                information.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 2

            # Estilo do cabeçalho: fundo colorido azul e letras brancas
            header_style = NamedStyle(name="header_style")
            header_style.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            header_style.font = Font(color="FFFFFF", bold=True)

            # Aplique o estilo de cabeçalho à primeira linha inteira
            information.row_dimensions[1].height = 20  # Defina a altura da linha
            for cell in information[1]:
                cell.style = header_style

            # Salve o arquivo Excel
            workbook.save(fr"C:\Users\lucas.sampaio\PACCAR Inc\Purchasing Brasil-DAF-Teams - Documentos\NPG - Brasil\2 - Macro\1 - Email\Informações para envio\RFX - {supplier}.xlsx")
            workbook.close()



class email:
    def create_list_email(self):
        book = openpyxl.load_workbook(r'C:\Users\lucas.sampaio\PACCAR Inc\Purchasing Brasil-DAF-Teams - Documentos\NPG - Brasil\2 - Macro\1 - Email\Macro - Email - Envio de cotação.xlsx',data_only=True)
        Dados_page = book['Emails']

        for row in Dados_page.iter_rows(min_row=3):
            fornecedor = row[1].value
            nome_fornecedor = row[2].value
            email = row[3].value
            key = {
                "Vendor":fornecedor,
                "Nome":nome_fornecedor,
                "Email":email
            }
            list_emails.append(key)

        print(list_emails)
    


    def create_email(self):

        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        user = input("What is your name? ")
        print()
        job = input("What is your function? ")

        for supplier in list_per_supplier:
            for email in list_emails:
                if supplier == email["Vendor"]:
                    email_cm = email["Email"]
            email = outlook.createitem(0)
            email.to = email_cm
            email.subject = f"Pending Modifications - Week {week_num}"
            email.htmlbody = f"""
            <p>Hello {supplier}, how are you?</p>
            <p></p>
            <p>I attached all the modifications that you have on status "pending".</p>
            <p>Can you contact the supplier to implement this modifications? My trigger to finish it is having the contract for Brazil on GPPS and/or having the current drawing revision commercial alligned.</p>
            <p>If you have any doubt, I will be happy to receive a teams message to talk about.</p>
            <p></p>
            <p>Best Regards</p>
            <p>{user} - {job}</p>
            """
            anexo = fr"C:\Users\lucas.sampaio\PACCAR Inc\Purchasing Brasil-DAF-Teams - Documentos\NPG - Brasil\2 - Macro\1 - Email\Informações para envio\RFX - {supplier}.xlsx"
            email.Attachments.Add(anexo)
            email.Display(True)


def main():
    List_manipulation = list()
    List_manipulation.list_of_pending()
    List_manipulation.list_per_cm()
    Email_manipulation = email()
    Email_manipulation.create_list_email()
    Email_manipulation.create_email()


if __name__ == "__main__":
    main()